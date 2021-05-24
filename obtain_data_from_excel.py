#Origional Author: dchidell.cisco.com
#Adaptions made by: scwhitin.cisco.com

from openpyxl import load_workbook

class excelReader():
    workbook_name = None
    heading_dict = {}
    workbook_data = []
    chassis_list = []
    status_dict = {}

    def __init__(self,workbook_name):
        self.workbook_name = workbook_name
        self.__get_worksheet_titles__()
        self.get_worksheet_data()

    def get_worksheet_data(self):
        self.workbook_data = {}
        workbook = load_workbook(self.workbook_name)
        sheet_titles = list(workbook.sheetnames)
        for sheet_name in sheet_titles:
            self.workbook_data[sheet_name] = []
            sheet = workbook[sheet_name]
            for idx,row in enumerate(sheet.rows):
                rowdict = {}
                if idx == 0: continue
                for colid,cell in enumerate(row):
                    data = cell.value
                    if data is None or data == '':
                        break
                    rowdict[self.heading_dict[sheet_name][colid]] = data
                if rowdict != {}:
                    self.workbook_data[sheet_name].append(rowdict)
        return self.workbook_data
        
    def __get_worksheet_titles__(self):
        workbook = load_workbook(self.workbook_name)
        sheet_titles = list(workbook.sheetnames)
        self.heading_dict = {}
        for sheet_name in sheet_titles:
            sheet = workbook[sheet_name]
            self.heading_dict[sheet_name] = []
            for idx,row in enumerate(sheet.rows):
                if idx > 0: break         
                for cell in row:
                    data = cell.value
                    if data is None or data == '':
                        break
                    self.heading_dict[sheet_name].append(data)
    
    def get_api_base_information(self):
        workbook = load_workbook(self.workbook_name)
        sheet = workbook["Base"]
        base_data_dict = {'Platform':sheet['B1'].value,'APIServerIP':sheet['B2'].value,'APIServerPort':sheet['B3'].value,'Username':sheet['B4'].value,'Password':sheet['B5'].value,
        'LicensingServerIP':sheet['B6'].value,'LicenseMode':sheet['B7'].value,'LicenseTier':sheet['B8'].value,'DebugMode':sheet['B9'].value,'ForcePortOwnership':sheet['B10'].value}
        return base_data_dict

    def obtain_chassis_ip(self):
        self.chassis_list = []
        for row in self.workbook_data['Physical']:
            #print(row)
            self.chassis_list.append(row['Chassis IP'])
        return self.chassis_list
    
    def obtain_status(self):
        self.status_dict = {}
        try:
            for row in self.workbook_data['Build_Information']:
                if row['Include'].lower() == 'yes': 
                    status = True
                else:
                    status = False
                self.status_dict[row['Worksheet']] = status
            return self.status_dict

        except Exception as e:
            print(e)
   

    

